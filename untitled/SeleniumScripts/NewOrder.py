__author__='SeanPark_ViaSat'

import unittest
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

import datetime
import time
import string
import random
import openpyxl
import pandas as pd
import xlsxwriter
import xlwt
import os

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import selenium.webdriver.support.ui as ui
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys


class NewOrder(unittest.TestCase):

    bankRoutingNumber = '122105278'
    bankAccountNumber = '0000000016'
    paymentType = ''

    def setUp(self):
        self.driver = webdriver.Chrome("C:\\Selenium\\chromedriver.exe")

        self.wb = openpyxl.load_workbook('NewConnectOrders.xlsx')

        self.wbAddress = openpyxl.load_workbook(('./Data/Addresses.xlsx'))
        self.sheetAddress = self.wbAddress['Sheet1']
        self.username = self.sheetAddress.cell(row=4, column=2).value
        self.password = self.sheetAddress.cell(row=4, column=3).value
        self.salesChannel = self.sheetAddress.cell(row=4, column=4).value
        self.customerType = self.sheetAddress.cell(row=4, column=5).value
        self.failures = []

    def test_firstTest(self):

        driver = self.driver

        driver.get("https://ordermgmt.test.exede.net/PublicGUI-SupportGUI/v1/pages/addcustomer/serviceAvailability.xhtml")

        print("driver.title : " + driver.title)

        self.assertIn("", driver.title)

        if driver.title != "Test":
            self.failures.append(driver.title + '!=' + "Test")

        driver.implicitly_wait(50)

        driver.set_page_load_timeout(30)

        driver.get("https://ordermgmt.test.exede.net/PublicGUI-SupportGUI/v1/pages/addcustomer/serviceAvailability.xhtml")

        driver.maximize_window()

        driver.implicitly_wait(20)

        driver.find_element_by_xpath("//*[@id=\"document:body\"]/table/tbody/tr[2]/td/form/table/tbody/tr[3]/td[2]/input").send_keys(self.username)

        driver.find_element_by_xpath(
            "//*[@id=\"document:body\"]/table/tbody/tr[2]/td/form/table/tbody/tr[4]/td[2]/input").send_keys(self.password)

        driver.find_element_by_name("submit").click()

        driver.implicitly_wait(5)

        self.addCustomerTab = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:add"]'))
        )

        self.addCustomerTab.click()

        for item in range(0, 1):

            driver.implicitly_wait(3)
            time.sleep(1)

            if self.salesChannel == 'WB_DIRECT':
                driver.find_element_by_xpath("//*[@id=\"addCustomerForm:salesChannelMenu\"]/option[2]").click()

            # salesChannelOption = selectSalesChannel(salesChannel)
            # select = Select(driver.find_element_by_id('addCustomerForm:salesChannelMenu'))
            #
            # select.select_by_visible_text(salesChannel).click()

            now = datetime.datetime.now()

            currentYear = str(now.year)

            months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']

            hexdigits = list(string.hexdigits)
            del hexdigits[10:16]

            # print(hexdigits)

            randomMac = "AA:BB:CC:"

            for x in range(0, 6):
                randomNumber = random.choice(hexdigits)
                randomMac = randomMac + randomNumber
                if x % 2 != 0 and len(randomMac) < 17:
                    randomMac = randomMac + ":"

            print("Mac Address : " + randomMac)

            self.randomMacNoColon = randomMac.replace(':', '')

            print(self.randomMacNoColon)

            currentMonth = months[now.month - 1]

            currentDay = ""

            if now.day < 10:
                currentDay = '0' + str(now.day)
            else:
                currentDay = str(now.day)

            transactionReference = "SPark_" + currentDay + currentMonth + currentYear + str(item)

            self.newSheet = currentDay + "-" + currentMonth + "-" + currentYear

            self.wb.create_sheet(self.newSheet)

            self.ws = self.wb[self.newSheet]

            self.ws.cell(row=1, column=2).value = 'Transaction Reference'
            self.ws.cell(row=1, column=3).value = 'Service Agreement'
            self.ws.cell(row=1, column=4).value = 'MAC'
            self.ws.cell(row=2, column=2).value = transactionReference
            self.ws.cell(row=2, column=4).value = randomMac

            driver.implicitly_wait(2)
            driver.find_element_by_xpath("//*[@id=\"addCustomerForm:transactionReference\"]").send_keys(
                transactionReference)

            # driver.implicitly_wait(2)
            # driver.find_element_by_xpath("//*[@id=\"addCustomerForm:namesIdName1\"]").send_keys("Spider")

            time.sleep(1)
            firstNameField = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID,
                                                'addCustomerForm:namesIdName1'))
            )

            firstNameField.send_keys("Spider")

            driver.implicitly_wait(2)

            time.sleep(1)
            lastNameField = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID,
                                                'addCustomerForm:namesIdName3'))
            )

            lastNameField.send_keys("Man")

            time.sleep(1)

            driver.implicitly_wait(2)
            driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableAddress1\"]").send_keys(
                "12017 E Lake Cir")
            time.sleep(1)

            driver.implicitly_wait(2)
            driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableCity\"]").send_keys("Englewood")
            time.sleep(1)

            driver.implicitly_wait(2)
            driver.find_element_by_xpath(
                "//*[@id=\"addCustomerForm:addressIdMaybeTableStateAddressState\"]/option[7]").click()
            time.sleep(1)

            driver.implicitly_wait(2)
            driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableZip\"]").send_keys("80111")
            time.sleep(1)

            driver.implicitly_wait(2)
            driver.find_element_by_xpath("//*[@id=\"addCustomerForm:primaryPhoneIdMaybeTablePhoneNumber\"]").send_keys(
                "7204823823")
            time.sleep(1)

            driver.implicitly_wait(2)
            driver.find_element_by_xpath("//*[@id=\"addCustomerForm:emailAddressId\"]").send_keys(
                "sean.park@viasat.com")
            time.sleep(1)

            driver.implicitly_wait(2)
            driver.find_element_by_xpath("//*[@id=\"addCustomerForm:Birthdate\"]").send_keys("12/15/1973")
            time.sleep(1)

            driver.implicitly_wait(2)
            driver.find_element_by_xpath("//*[@id=\"addCustomerForm:nextButtonId\"]").click()
            time.sleep(1)

            # Contacts Page

            # creditCheckPassed = WebDriverWait(driver, 60).until(
            #     EC.presence_of_element_located((By.XPATH, '// *[@id = "addCustomerForm:_id93"]/tbody/tr/td/span'))
            # )

            # print("creditCheckPassed? : " + creditCheckPassed.text)

            time.sleep(5)

            assert " " in driver.page_source

            driver.implicitly_wait(2)

            time.sleep(1)

            customerReferenceField = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID,
                                                'addCustomerForm:customerReference'))
            )

            customerReferenceField.send_keys(transactionReference)

            driver.implicitly_wait(2)

            time.sleep(1)
            accountReferenceField = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID,
                                                'addCustomerForm:accountReference'))
            )

            accountReferenceField.send_keys(transactionReference)

            driver.implicitly_wait(2)
            time.sleep(1)
            driver.find_element_by_xpath("//*[@id=\"addCustomerForm:nextButtonId\"]").click()

            # Packages Page

            packagesTitle = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:packagesHeaderLabel"]'))
            )

            # Sometimes, radio button is not checked by default. So intentionally click it.
            # id is dynamically created. id114 or id112
            # driver.find_element_by_xpath('//*[@id="addCustomerForm:_id114:_2"]').click()
            # driver.find_element_by_xpath('//*[@id="addCustomerForm:_id112:_2"]').click()

            time.sleep(1)

            packageRadioButtonFirst = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, '//input[starts-with(@value, "$")]'))
            )

            # "is not clickable at point" error. Another element is covering the element to click. I could use execute_script() to click on this.
            nextButton = driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]')
            driver.execute_script("arguments[0].click();", nextButton)

            driver.implicitly_wait(2)
            time.sleep(1)

            # Options Page
            optionsTitle = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:optionsLabel"]'))
            )

            print("optionsTitle : " + optionsTitle.text)

            driver.implicitly_wait(2)
            time.sleep(1)
            driver.find_element_by_xpath("//*[@id=\"addCustomerForm:_1selectionPackages:_1\"]").click()

            driver.implicitly_wait(2)
            time.sleep(1)
            # driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]').click()

            optionsPageNextButton = driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]')
            driver.execute_script("arguments[0].click();", optionsPageNextButton)

            driver.implicitly_wait(2)
            time.sleep(1)

            # Payment Page
            paymentMethodTitle = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:recurringPaymentInfoLabel"]'))
            )

            driver.implicitly_wait(2)
            time.sleep(1)
            driver.find_element_by_xpath(
                "//*[@id=\"addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdcreditCardTypeId\"]/option[3]").click()

            driver.implicitly_wait(2)
            time.sleep(1)
            driver.find_element_by_xpath(
                "//*[@id=\"addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdNumberId\"]").send_keys(
                "4012000077777777")

            driver.implicitly_wait(2)
            time.sleep(1)
            driver.find_element_by_xpath(
                "//*[@id=\"addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdExpireMonthIdMonthId\"]/option[5]").click()

            driver.implicitly_wait(2)
            time.sleep(1)
            driver.find_element_by_xpath(
                '//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdExpireYearIdYearId"]/option[3]').click()

            driver.implicitly_wait(2)
            time.sleep(1)
            driver.find_element_by_xpath(
                '//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdFirstNameId"]').send_keys(
                "VISA")

            driver.implicitly_wait(2)
            time.sleep(1)
            driver.find_element_by_xpath(
                '//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdLastNameId"]').send_keys(
                "APPROVAL")

            driver.implicitly_wait(2)
            time.sleep(1)
            driver.find_element_by_xpath(
                '//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdAddressZip"]').send_keys(
                "80111")

            driver.implicitly_wait(2)
            time.sleep(1)
            driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]').click()

            # Review Page

            self.scheduleButton = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:scheduleInstallationButtonId"]'))
            )

            self.scheduleButton.click()
            time.sleep(1)

            # Schedule Page

            self.submitOrderButton = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:submitButtonId"]'))
            )

            self.submitOrderButton.click()
            time.sleep(0.3)

            # wait for order reference number created

            # Confirmation Page

            printButton = WebDriverWait(driver, 180).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:printButtonId"]'))
            )

            self.serviceAgreementReference = driver.find_element_by_xpath(
                '//*[@id="addCustomerForm:serviceAgreementReference"]').text

            print("Sales Channel : " + self.salesChannel)
            print("External Account Reference : " + self.serviceAgreementReference)

            self.newOrderButton = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:newOrderButtonId"]'))
            )

            self.newOrderButton.click()

            driver.implicitly_wait(20)
            time.sleep(1)

            self.transactionInfoTitle = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:transactionInfoLabel"]'))
            )

            print("transactionInfoTitle : " + self.transactionInfoTitle.text)
            driver.get('https://spyglass01.test.wdc1.wildblue.net:8443/SpyGlass/')

            self.referenceType = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,
                                                '/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[1]/select/option[5]'))
            )

            self.referenceType.click()

            self.referenceValue = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,
                                                '/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[2]/input'))
            )

            self.referenceValue.send_keys(transactionReference)

            self.externalSystem = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,
                                                '/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[3]/div/select/option[2]'))
            )

            self.externalSystem.click()

            driver.implicitly_wait(20)
            time.sleep(1)

            driver.find_element_by_xpath(
                '/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[4]/input[1]').click()

            self.fsmCustomerCode = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="datatable"]/tbody/tr[1]/td[1]/div[1]'))
            )

            self.serviceAgreementNumber = driver.find_element_by_xpath(
                '//*[@id="data"]/table[1]/tbody/tr[2]/td/table/tbody/tr[2]/td[12]').text

            print('serviceAgreementNumber : ' + self.serviceAgreementNumber)

            driver.save_screenshot('./Reports/' + self.serviceAgreementNumber + '.png')

            driver.implicitly_wait(20)
            time.sleep(1)

            self.ws.cell(row=2, column=3).value = self.serviceAgreementNumber

            #######################################
            ###
            ### Provisioning Starting from here....
            ###
            #######################################

            # driver = webdriver.Chrome("C:\\Selenium\\chromedriver.exe")
            # driver = webdriver.Ie("C:\\Selenium\\IEDriverServer.exe")

            # driver.set_page_load_timeout(30)

            time.sleep(6)

            # driver.get("https://ordermgmt.test.exede.net/PublicGUI-SupportGUI/v1/pages/addcustomer/serviceAvailability.xhtml")
            #
            # driver.get('https://spyglass01.test.wdc1.wildblue.net:8443/SpyGlass/')

            self.installGUI = "https://igui-installationgui.test.wdc1.wildblue.net/InternalGUI-InstallationGUI/"

            self.installGUIwithMac = self.installGUI + "?n=" + self.randomMacNoColon

            driver.get(self.installGUIwithMac)

            # serviceAgreementNumber = '402907978'

            self.screenshotDirectory = './Reports/' + self.serviceAgreementNumber + '_' + driver.name

            if not os.path.exists(self.screenshotDirectory):
                os.makedirs(self.screenshotDirectory)

            driver.maximize_window()

            time.sleep(3)

            print("Web Browser in test : " + driver.name)

            ### if it's IE, it needs to bypass security warning
            if driver.name == "internet explorer":
                print("driver is IE")
                self.continueLink = driver.find_element_by_id('overridelink')
                self.continueLink.click()

            driver.implicitly_wait(20)

            self.activationKeyField = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:activationKey"]'))
            )

            self.activationKeyField.send_keys(self.serviceAgreementNumber)

            self.installButton = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id36"]'))
            )

            driver.save_screenshot(self.screenshotDirectory + '/1_welcomeToServiceActivation.png')

            self.installButton.click()

            time.sleep(2)

            self.installerNumberField = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:installerId"]'))
            )

            self.installerNumberField.send_keys("99072761")

            driver.save_screenshot(self.screenshotDirectory + '/2_customerConfirmationNewInstallation.png')

            self.continueInstallButton = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id53"]'))
            )

            self.continueInstallButton.click()

            time.sleep(5)

            self.emailConfirmationButton = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id30"]'))
            )

            driver.save_screenshot(self.screenshotDirectory + '/3_emailConfirmationAndUpdate.png')

            self.emailConfirmationButton.click()

            time.sleep(10)

            # qOIcontinueButton = WebDriverWait(driver, 60).until(
            #     EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id50"]'))
            # )

            self.thankYouTag = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id40"]'))
            )

            print('thankYouTag Text : ' + self.thankYouTag.text)

            self.qOIcontinueButton = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//input[@type="submit"]'))
            )

            driver.save_screenshot(self.screenshotDirectory + '/4_qualityOfInstall.png')

            self.qOIcontinueButton.click()

            time.sleep(6)

            self.customerButton = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, '//input[@value="Customer"]'))
            )

            self.customerButton.click()

            driver.save_screenshot(self.screenshotDirectory + '/5_newCustomerAccountSetup.png')

            self.lastFourField = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:paymentAuthentication"]'))
            )

            self.lastFourField.send_keys("7777")

            self.ccContinueButton = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//input[@value="Continue"]'))
            )

            self.ccContinueButton.click()

            time.sleep(5)

            # driver.switch_to_frame(1)

            self.pdfIFrame = driver.find_element_by_xpath('//*[@id="installerForm:j_id20"]/iframe')

            # print(pdfIFrame.get_attribute('src'))

            driver.switch_to_default_content()

            driver.switch_to_frame(self.pdfIFrame)

            time.sleep(5)

            driver.save_screenshot(self.screenshotDirectory + '/6_customerAgreement.png')

            time.sleep(2)

            self.getStartedButton = WebDriverWait(driver, 60).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="pnlElectronic"]/div/div[1]/button[1]/i'))
            )

            print("getStartedButtonAttribute : " + self.getStartedButton.get_attribute('class'))

            self.getStartedButton.click()

            time.sleep(3)

            self.signField = driver.find_element_by_xpath('//*[@id="location1"]/div[2]/div[1]/input')

            print(self.signField.get_attribute('type'))

            self.signField.send_keys("Spider Man")

            time.sleep(3)

            driver.save_screenshot(self.screenshotDirectory + '/7_customerAgreementAfterSign.png')

            self.finishSubmitButton = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="completePopupContainer"]/div/div[1]/button'))
            )

            self.finishSubmitButton.click()

            time.sleep(2)

            driver.save_screenshot(self.screenshotDirectory + '/8_eSignSubmitted.png')

            time.sleep(2)

            driver.switch_to_default_content()

            self.continueButtonAfterSign = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id25"]'))
            )

            print('continueButtonAfterSign attribute : ' + self.continueButtonAfterSign.get_attribute('class'))

            driver.save_screenshot(self.screenshotDirectory + '/9_eSignComplete.png')

            self.continueButtonAfterSign.click()

            time.sleep(3)

            driver.save_screenshot(self.screenshotDirectory + '/10_activatingModem.png')

            self.activatingModemContinueButton = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id35"]'))
            )

            self.activatingModemContinueButton.click()

            time.sleep(2)

            self.confirmationMessage = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id19"]'))
            )

            driver.save_screenshot(self.screenshotDirectory + '/11_confirmation.png')

            if "Success!" not in driver.page_source:
                self.failures.append("Success! not in " + driver.page_source)

            # assert(self.failures == [], str(self.failures))

    def teatDown(self):

        self.wb.save('NewConnectOrders.xlsx')

        self.driver.close()

if __name__ == "__main__":
    unittest.main()




