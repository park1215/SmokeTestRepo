__author__='SeanPark_ViaSat'

import datetime
import time
import string
import random
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import pandas as pd
import xlsxwriter
import xlwt
import os as winos

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import selenium.webdriver.support.ui as ui
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
import logging

bankRoutingNumber = '122105278'

bankAccountNumber = '0000000016'

wb = openpyxl.load_workbook('./Reports/NewConnectOrders.xlsx')

wbAddress = openpyxl.load_workbook('./Data/Addresses.xlsx')

sheetAddress = wbAddress['Sheet1']

rowLength = sheetAddress.max_row

columnLength = sheetAddress.max_column

currentRow = 2

##################################
### This code block is for logging
##################################
logdate = datetime.datetime.now().strftime('%Y-%m-%d')
logFileName = "NewConnect_Provisioning_" + logdate + '.log'

logger = logging.getLogger('NewConnect_Provisioing')
logger.setLevel(logging.DEBUG)
# create file handler which logs even debug messages
fh = logging.FileHandler('./Reports/'+logFileName)
fh.setLevel(logging.DEBUG)
# create console handler with a higher log level
ch = logging.StreamHandler()
ch.setLevel(logging.DEBUG)
# create formatter and add it to the handlers
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
ch.setFormatter(formatter)
fh.setFormatter(formatter)
# add the handlers to logger
logger.addHandler(ch)
logger.addHandler(fh)

logger.debug('-----------------------------------------------------------------------')
logger.debug('-----------------------------------------------------------------------')
# logger.info('info message')
# logger.warn('warn message')
# logger.error('error message')
# logger.critical('critical message')

for item in range(currentRow, rowLength+1):
    logger.debug("Number of Orders to be created in this run: " + str(rowLength-1))

    logger.debug("Current Row in the data table : " + str(currentRow-1))

    username = sheetAddress.cell(row=currentRow, column=1).value

    password = sheetAddress.cell(row=currentRow, column=2).value

    salesChannel = sheetAddress.cell(row=currentRow, column=3).value

    customerType = sheetAddress.cell(row=currentRow, column=4).value

    addressLine1 = sheetAddress.cell(row=currentRow, column=7).value

    city = sheetAddress.cell(row=currentRow, column=9).value

    state = sheetAddress.cell(row=currentRow, column=10).value

    zipCode = sheetAddress.cell(row=currentRow, column=11).value

    voipIncluded = sheetAddress.cell(row=currentRow, column=17).value

    if voipIncluded.lower() == "no":
        voipIncluded = False

    elif voipIncluded.lower() == "yes":
        voipIncluded = True
    else:
        voipIncluded = None

    logger.debug("voip is included in the data table? : " + str(voipIncluded))

    paymentType = sheetAddress.cell(row=currentRow, column=18).value

    packageName = sheetAddress.cell(row=currentRow, column=25).value

    satelliteName = sheetAddress.cell(row=currentRow, column=26).value

    logger.debug("Payment type in the data table: " + paymentType)

    logger.debug("Package Name : " + packageName)

    logger.debug("Satellite Name : " + satelliteName)

    driver = webdriver.Chrome("C:\\Selenium\\chromedriver.exe")
    # driver = webdriver.Ie("C:\Selenium\IEDriverServer.exe");

    driver.implicitly_wait(50)

    driver.set_page_load_timeout(30)

    driver.get("https://ordermgmt.test.exede.net/PublicGUI-SupportGUI/v1/pages/addcustomer/serviceAvailability.xhtml")

    driver.maximize_window()

    driver.implicitly_wait(20)

    driver.find_element_by_xpath(
        "//*[@id=\"document:body\"]/table/tbody/tr[2]/td/form/table/tbody/tr[3]/td[2]/input").send_keys(username)

    driver.find_element_by_xpath(
        "//*[@id=\"document:body\"]/table/tbody/tr[2]/td/form/table/tbody/tr[4]/td[2]/input").send_keys(password)

    driver.find_element_by_name("submit").click()

    driver.implicitly_wait(5)

    # addCustomerTab has id of "addCustomerForm:add" OR, sometimes "add".
    # so I decided to use href, which could be more reliable
    addCustomerTab = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//a[@href="/PublicGUI-SupportGUI/v1/pages/addcustomer/serviceAvailability.xhtml"]'))
    )

    addCustomerTab.click()

    driver.implicitly_wait(3)
    time.sleep(1)

    if salesChannel.lower() == 'WB_DIRECT'.lower():
        # driver.find_element_by_xpath("//*[@id=\"addCustomerForm:salesChannelMenu\"]/option[2]").click()
        salesChannelSelected = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:salesChannelMenu"]/option[2]'))
        )
    salesChannelSelected.click()

    if customerType.lower() == 'Residential'.lower():
        # driver.find_element_by_xpath("//*[@id=\"addCustomerForm:salesChannelMenu\"]/option[2]").click()
        customerTypeSelected = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:customerTypeMenu"]/option[1]'))
        )

    customerTypeSelected.click()

    # salesChannelOption = selectSalesChannel(salesChannel)
    # select = Select(driver.find_element_by_id('addCustomerForm:salesChannelMenu'))
    #
    # select.select_by_visible_text(salesChannel).click()

    now = datetime.datetime.now()

    currentYear = str(now.year)

    months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']

    hexdigits = list(string.hexdigits)
    del hexdigits[10:16]

    # logger.debug(hexdigits)

    randomMac = "AA:BB:CC:"

    for x in range(0, 6):
        randomNumber = random.choice(hexdigits)
        randomMac = randomMac + randomNumber
        if x % 2 != 0 and len(randomMac) < 17:
            randomMac = randomMac + ":"

    logger.debug("Mac Address : " + randomMac)

    randomMacNoColon = randomMac.replace(':', '')

    logger.debug(randomMacNoColon)

    currentMonth = months[now.month - 1]

    currentDay = ""

    if now.day < 10:
        currentDay = '0' + str(now.day)
    else:
        currentDay = str(now.day)

    # transactionReference = "SPark_" + currentDay + currentMonth + currentYear + str(item+15)

    transactionReference = "SPark_" + str(randomMacNoColon)

    SupportPortalScreenshotDirectory = './Reports/' + randomMacNoColon

    if not winos.path.exists(SupportPortalScreenshotDirectory):
        winos.makedirs(SupportPortalScreenshotDirectory)

    newSheetName = currentDay + "-" + currentMonth + "-" + currentYear

    sheetList = wb.sheetnames

    if newSheetName not in sheetList:
        logger.debug(newSheetName + " Not exist. It creates a new sheet : " + newSheetName)
        wb.create_sheet(newSheetName)

    ws = wb[newSheetName]

    end_of_sheet = ws.max_row

    start_of_sheet = end_of_sheet + 1

    logger.debug("Start row of report sheet : " + str(start_of_sheet))

    ws.cell(row=1, column=2).value = 'Transaction Reference'
    ws.cell(row=1, column=3).value = 'Service Agreement'
    ws.cell(row=1, column=4).value = 'MAC'
    ws.cell(row=1, column=5).value = 'Sales Channel'
    ws.cell(row=1, column=6).value = 'Customer Type'
    ws.cell(row=1, column=7).value = 'Payment Type'
    ws.cell(row=1, column=8).value = 'Satellite'
    ws.cell(row=1, column=9).value = 'Package'
    ws.cell(row=1, column=10).value = 'Acct. Reference'
    ws.cell(row=1, column=11).value = 'VoipIncluded'
    ws.cell(row=1, column=12).value = 'Beam'

    ws.cell(row=start_of_sheet, column=2).value = transactionReference
    ws.cell(row=start_of_sheet, column=4).value = randomMac
    ws.cell(row=start_of_sheet, column=5).value = salesChannelSelected.text
    ws.cell(row=start_of_sheet, column=6).value = customerTypeSelected.text
    ws.cell(row=start_of_sheet, column=11).value = sheetAddress.cell(row=currentRow, column=17).value

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:transactionReference\"]").send_keys(transactionReference)

    # driver.implicitly_wait(2)
    # driver.find_element_by_xpath("//*[@id=\"addCustomerForm:namesIdName1\"]").send_keys("Spider")

    time.sleep(1)
    driver.implicitly_wait(2)
    firstNameField = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID,
                                        'addCustomerForm:namesIdName1'))
    )

    firstNameField.send_keys("Spider")

    driver.implicitly_wait(2)

    time.sleep(1)
    lastNameField = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID,
                                        'addCustomerForm:namesIdName3'))
    )

    lastNameField.send_keys("Man")

    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableAddress1\"]").send_keys(addressLine1)
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableCity\"]").send_keys(city)
    time.sleep(1)

    driver.implicitly_wait(2)
    stateAddress = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH,
                                        '//*[@id="addCustomerForm:addressIdMaybeTableStateAddressState"]/option[7]'))
    )
    stateAddress.click()
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableZip\"]").send_keys(zipCode)
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:primaryPhoneIdMaybeTablePhoneNumber\"]").send_keys("7204823823")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:emailAddressId\"]").send_keys("sean.park@viasat.com")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:Birthdate\"]").send_keys("01/01/1980")
    time.sleep(1)

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/1_serviceability.png')

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:nextButtonId\"]").click()
    time.sleep(1)

    # Contacts Page

    # creditCheckPassed = WebDriverWait(driver, 60).until(
    # EC.presence_of_element_located((By.XPATH, '// *[@id = "addCustomerForm:_id91"]/tbody/tr/td/span'))
    # )

    driver.implicitly_wait(2)

    # assert "The Credit Check passed." in driver.page_source

    time.sleep(1)

    customerReferenceField = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID,
                                        'addCustomerForm:customerReference'))
    )

    customerReferenceField.send_keys(transactionReference)

    driver.implicitly_wait(2)
    time.sleep(1)
    accountReferenceField = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID,
                                        'addCustomerForm:accountReference'))
    )

    accountReferenceField.send_keys(transactionReference)

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/2_contacts.png')

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:nextButtonId\"]").click()

    # Packages Page

    packagesTitle = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:packagesHeaderLabel"]'))
    )

    # Sometimes, radio button is not checked by default. So intentionally click it.
    # id is dynamically created. id114 or id112
    # driver.find_element_by_xpath('//*[@id="addCustomerForm:_id114:_2"]').click()
    # driver.find_element_by_xpath('//*[@id="addCustomerForm:_id112:_2"]').click()

    time.sleep(1)

    # packageRadioButtonFirst = WebDriverWait(driver, 20).until(
    #     EC.presence_of_element_located((By.XPATH, '//input[starts-with(@value, "$")]'))
    # )

    # This statement is to be replaced with dictionary type switch statement.
    # if packageName.lower() == 'Unlimited Bronze 12'.lower():
    #     packageRadioButton = WebDriverWait(driver, 20).until(
    #         EC.presence_of_element_located((By.XPATH, '//input[starts-with(@value, "$")]'))
    #     )

    if packageName.lower() == 'Unlimited Bronze 12'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_0"]'))
        )

    elif packageName.lower() == 'Unlimited Silver 12'.lower() and satelliteName.lower() == 'VS1'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_1"]'))
        )

    elif packageName.lower() == 'Unlimited Gold 12'.lower() and satelliteName.lower() == 'VS1'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_2"]'))
        )

    elif packageName.lower() == 'Unlimited Silver 25'.lower() and satelliteName.lower() == 'VS1'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_4"]'))
        )

    elif packageName.lower() == 'Unlimited Gold 30'.lower() and satelliteName.lower() == 'VS1'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_5"]'))
        )

    elif packageName.lower() == 'Unlimited Silver 25'.lower() and satelliteName.lower() == 'VS2'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_1"]'))
        )

    elif packageName.lower() == 'Unlimited Gold 50'.lower() and satelliteName.lower() == 'VS2'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_2"]'))
        )
    packageRadioButton.click()

    time.sleep(1)

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/3_packages.png')

    # "is not clickable at point" error. Another element is covering the element to click. I could use execute_script() to click on this.
    nextButton = driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]')
    driver.execute_script("arguments[0].click();", nextButton)

    driver.implicitly_wait(2)
    time.sleep(1)

    # Options Page
    optionsTitle = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:optionsLabel"]'))
    )

    logger.debug("Options Title displayed is : " + optionsTitle.text)

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:_1selectionPackages:_1\"]").click()

    driver.implicitly_wait(2)
    time.sleep(1)

    #  This is to add voip service.
    # voipSelectButton = WebDriverWait(driver, 180).until(
    #     EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:_id248"]'))
    # )
    #
    # voipSelectButton.click()

    # Checking EasyCare option
    driver.find_element_by_xpath('//input[@type="checkbox"]').click()

    driver.implicitly_wait(2)
    time.sleep(1)
    # driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]').click()

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/4_options.png')

    optionsPageNextButton = driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]')
    driver.execute_script("arguments[0].click();", optionsPageNextButton)

    driver.implicitly_wait(2)
    time.sleep(1)

    # VoIP Page
    #
    # driver.implicitly_wait(2)
    # time.sleep(1)
    #
    # voipPassword = "Qkrtmd!1"
    #
    # voipPasswordField = WebDriverWait(driver, 180).until(
    #     EC.presence_of_element_located((By.ID, 'addCustomerForm:password'))
    # )
    #
    # voipPasswordField.send_keys(voipPassword)
    #
    # driver.implicitly_wait(2)
    # time.sleep(1)
    #
    # logger.debug("voipUserName : " + driver.find_element_by_xpath('//*[@id="addCustomerForm:userName"]').text)
    #
    # voipPasswordConfirmField = WebDriverWait(driver, 180).until(
    #     EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:confirm"]'))
    # )
    #
    # voipPasswordConfirmField.send_keys(voipPassword)
    #
    # driver.implicitly_wait(2)
    # time.sleep(1)
    #
    # streetAddressNumberField = WebDriverWait(driver, 180).until(
    #     EC.presence_of_element_located((By.ID, 'addCustomerForm:streetNumber'))
    # )
    #
    # streetAddressNumberField.send_keys("12017")
    #
    # driver.implicitly_wait(2)
    # time.sleep(1)
    #
    # streetAddressStreetNameField = WebDriverWait(driver, 180).until(
    #     EC.presence_of_element_located((By.ID, 'addCustomerForm:streetName'))
    # )
    #
    # streetAddressStreetNameField.send_keys("E Lake Cir")
    #
    # driver.implicitly_wait(2)
    # time.sleep(1)
    #
    # phoneNumberSelected = WebDriverWait(driver, 180).until(
    #     EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:voipChooseNumberMenu"]/option[2]'))
    # )
    #
    # phoneNumberSelected.click()
    #
    # voipNextButton = driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]')
    #
    # driver.save_screenshot(SupportPortalScreenshotDirectory + '/5_voip.png')
    #
    # driver.execute_script("arguments[0].click();", voipNextButton)

    # Payment Page
    driver.implicitly_wait(2)
    time.sleep(5)

    try:
        paymentMethodTitle = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:recurringPaymentInfoLabel"]'))
        )
    except:
        logger.debug("payment method title is not displayed.")

    if paymentType.lower() == 'CC'.lower():
        driver.implicitly_wait(2)
        time.sleep(1)

        paymentTypeSelected = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,
                                            '//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableselectPaymentTypeChoiceId"]/option[1]'))
        )

        ws.cell(row=start_of_sheet, column=7).value = paymentTypeSelected.text

        driver.find_element_by_xpath("//*[@id=\"addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdcreditCardTypeId\"]/option[3]").click()

        driver.implicitly_wait(2)
        time.sleep(1)
        driver.find_element_by_xpath("//*[@id=\"addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdNumberId\"]").send_keys("4012000077777777")

        driver.implicitly_wait(2)
        time.sleep(1)
        driver.find_element_by_xpath("//*[@id=\"addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdExpireMonthIdMonthId\"]/option[5]").click()

        driver.implicitly_wait(2)
        time.sleep(1)
        driver.find_element_by_xpath('//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdExpireYearIdYearId"]/option[3]').click()

        driver.implicitly_wait(2)
        time.sleep(1)
        driver.find_element_by_xpath('//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdFirstNameId"]').send_keys("VISA")

        driver.implicitly_wait(2)
        time.sleep(1)
        driver.find_element_by_xpath('//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdLastNameId"]').send_keys("APPROVAL")

        driver.implicitly_wait(2)
        time.sleep(1)
        driver.find_element_by_xpath('//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdAddressZip"]').send_keys("80111")

    elif paymentType == 'EFT':
        driver.implicitly_wait(2)
        time.sleep(1)

        paymentTypeSelected = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH,
                                            '//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableselectPaymentTypeChoiceId"]/option[2]'))
        )

        ws.cell(row=start_of_sheet, column=7).value = paymentTypeSelected.text

        driver.find_element_by_xpath(
            '//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableselectPaymentTypeChoiceId"]/option[2]').click()

        driver.implicitly_wait(2)
        time.sleep(1)
        driver.find_element_by_xpath(
            '//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableEFTTypeIdBankRoutingNumberId"]').send_keys(bankRoutingNumber)

        driver.implicitly_wait(2)
        time.sleep(1)
        driver.find_element_by_xpath(
            '//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableEFTTypeIdBankAccountNumberId"]').send_keys(
            bankAccountNumber)

        driver.implicitly_wait(2)
        time.sleep(1)
        driver.find_element_by_xpath(
            '//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableEFTTypeIdFirstNameId"]').send_keys('EFT')

        driver.implicitly_wait(2)
        time.sleep(1)
        driver.find_element_by_xpath(
            '//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableEFTTypeIdLastNameId"]').send_keys(
            'APPROVAL')

        driver.implicitly_wait(2)
        time.sleep(1)
        driver.find_element_by_xpath(
            '//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableEFTTypeIdBusinessNameId"]').send_keys(
            "Business Name")
    try:
        taxJurisdictionDropdown = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:taxJurisdictionMenu"]'))
        )

        driver.find_element_by_xpath('//*[@id="addCustomerForm:taxJurisdictionMenu"]/option[2]').click()

        logger.debug("taxJurisdiction is a dropdown menu")
    except:
        logger.debug("taxJurisdiction is NOT a dropdown menu")

    # dropdownPresent = driver.find_element_by_xpath('//*[@id="addCustomerForm:taxJurisdictionMenu"]/option[2]').is_displayed()

    # if dropdownPresent:
    #     driver.find_element_by_xpath('//*[@id="addCustomerForm:taxJurisdictionMenu"]/option[2]').click()

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/5_payment.png')

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]').click()

    # Review Page

    scheduleButton = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:scheduleInstallationButtonId"]'))
    )

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/6_review.png')

    scheduleButton.click()
    time.sleep(2)

    # Schedule Page

    submitOrderButton = WebDriverWait(driver, 50).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:submitButtonId"]'))
    )

    time.sleep(1)

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/7_schedule.png')

    submitOrderButton.click()

    # wait for order reference number created

    # Confirmation Page

    try:
        printButton = WebDriverWait(driver, 180).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:printButtonId"]'))
        )
    except:
        logger.debug('FSM screen is not displayed. It continues to the next row.')
        continue

    serviceAgreementReference = driver.find_element_by_xpath('//*[@id="addCustomerForm:serviceAgreementReference"]').text

    logger.debug("Sales Channel for this order : " + salesChannel)
    logger.debug("External Account Reference for this order : " + serviceAgreementReference)

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/8_confirmation.png')

    satelliteUsed = driver.find_element_by_xpath(
        '//*[@id="addCustomerForm:satelliteId"]')

    ws.cell(row=start_of_sheet, column=8).value = satelliteUsed.text

    serviceSelected = driver.find_element_by_xpath(
        '//*[@id="addCustomerForm:ServicesTableGrid-2-0"]')

    ws.cell(row=start_of_sheet, column=9).value = serviceSelected.text

    accountReference = driver.find_element_by_xpath(
        '//*[@id="addCustomerForm:accountReference"]')

    ws.cell(row=start_of_sheet, column=10).value = accountReference.text

    beamSelected = driver.find_element_by_xpath(
        '//*[@id="addCustomerForm:beamId"]')

    ws.cell(row=start_of_sheet, column=12).value = beamSelected.text

    newOrderButton = driver.find_element_by_xpath('//*[@id="addCustomerForm:newOrderButtonId"]')

    newOrderButton.click()

    driver.implicitly_wait(20)
    time.sleep(1)

    transactionInfoTitle = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:transactionInfoLabel"]'))
    )

    ###########################################################
    ### Checking out the service agreement number from Spyglass
    ###########################################################
    driver.get('https://spyglass01.test.wdc1.wildblue.net:8443/SpyGlass/')

    referenceType = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[1]/select/option[5]'))
        )

    referenceType.click()

    referenceValue = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[2]/input'))
        )

    referenceValue.send_keys(transactionReference)

    externalSystem = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[3]/div/select/option[2]'))
        )

    externalSystem.click()

    driver.implicitly_wait(20)
    time.sleep(1)

    driver.find_element_by_xpath('/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[4]/input[1]').click()

    fsmCustomerCode = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="datatable"]/tbody/tr[1]/td[1]/div[1]'))
        )

    serviceAgreementNumber = driver.find_element_by_xpath('//*[@id="data"]/table[1]/tbody/tr[2]/td/table/tbody/tr[2]/td[12]').text

    logger.debug('serviceAgreementNumber for this order : ' + serviceAgreementNumber)

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/9_spyglass.png')

    driver.implicitly_wait(20)
    time.sleep(1)

    ws.cell(row=start_of_sheet, column=3).value = serviceAgreementNumber

    #######################################
    ###
    ### Provisioning Starting from here....
    ###
    #######################################

    logger.debug('Provisioning starts....')
    # driver = webdriver.Chrome("C:\\Selenium\\chromedriver.exe")
    # driver = webdriver.Ie("C:\\Selenium\\IEDriverServer.exe")

    # driver.set_page_load_timeout(30)

    time.sleep(6)

    # driver.get("https://ordermgmt.test.exede.net/PublicGUI-SupportGUI/v1/pages/addcustomer/serviceAvailability.xhtml")
    #
    # driver.get('https://spyglass01.test.wdc1.wildblue.net:8443/SpyGlass/')

    installGUI = "https://igui-installationgui.test.wdc1.wildblue.net/InternalGUI-InstallationGUI/"

    installGUIwithMac = installGUI + "?n=" + randomMacNoColon

    driver.get(installGUIwithMac)

    # serviceAgreementNumber = '402907978'

    provioningScreenshotDirectory = SupportPortalScreenshotDirectory + '/' + serviceAgreementNumber

    if not winos.path.exists(provioningScreenshotDirectory):
        winos.makedirs(provioningScreenshotDirectory)

    driver.maximize_window()

    time.sleep(3)

    logger.debug("Web Browser used in this test : " + driver.name)

    ### if it's IE, it needs to bypass security warning
    if driver.name == "internet explorer":
        continueLink = driver.find_element_by_id('overridelink')
        continueLink.click()

    driver.implicitly_wait(20)

    activationKeyField = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:activationKey"]'))
    )

    activationKeyField.send_keys(serviceAgreementNumber)

    installButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id36"]'))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/1_welcomeToServiceActivation.png')

    installButton.click()

    time.sleep(2)

    ### Customer confirmation New Installation Page

    installerNumberField = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:installerId"]'))
    )

    installerNumberField.send_keys("99072761")

    driver.save_screenshot(provioningScreenshotDirectory + '/2_customerConfirmationNewInstallation.png')

    continueInstallButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id53"]'))
    )

    continueInstallButton.click()

    time.sleep(5)

    emailConfirmationButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id30"]'))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/3_emailConfirmationAndUpdate.png')

    emailConfirmationButton.click()

    time.sleep(10)

    ### Quality of Install Page
    # qOIcontinueButton = WebDriverWait(driver, 60).until(
    #     EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id50"]'))
    # )

    thankYouTag = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id40"]'))
    )

    logger.debug('QOI in progress.....')

    qOIcontinueButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//input[@type="submit"]'))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/4_qualityOfInstall.png')

    qOIcontinueButton.click()

    ### Exede Voice Page
    # logger.debug('Entering Exede Voice if voip added to service...')
    #
    # voiceActivationPortalButton = WebDriverWait(driver, 60).until(
    #     EC.presence_of_element_located((By.XPATH, '//*[@value="Voice Activation Portal"]'))
    # )
    #
    # driver.save_screenshot(provioningScreenshotDirectory + '/5_voipActivation1.png')
    #
    # voiceActivationPortalButton.click()
    #
    # ### Exede Voice - Part 1: Complete teh Voice Activation Portal Process Page
    # ### Step #1 - Identify Account Page
    # time.sleep(2)
    # voipIFrame = driver.find_element_by_xpath('//*[@id="installerForm:j_id25"]/iframe')
    #
    # time.sleep(2)
    #
    # driver.switch_to_default_content()
    #
    # logger.debug("default content iFrame driver title : " + driver.title)
    #
    # driver.switch_to_frame(voipIFrame)
    #
    # logger.debug("VoiP activation iFrame entered...")
    #
    # logger.debug("VoiP iFrame driver title : " + driver.title)
    #
    # voipAccountNumberField = WebDriverWait(driver, 60).until(
    #     EC.presence_of_element_located((By.ID, 'inputAccountNumber'))
    # )
    #
    # voipAccountNumberField.send_keys(serviceAgreementNumber)
    #
    # voipLastNameField = WebDriverWait(driver, 60).until(
    #     EC.presence_of_element_located((By.ID, 'inputLastName'))
    # )
    #
    # voipLastNameField.send_keys("Man")
    #
    # voipIdentifyButton = WebDriverWait(driver, 60).until(
    #     EC.presence_of_element_located((By.XPATH, '//*[@type="submit"]'))
    # )
    #
    # driver.save_screenshot(provioningScreenshotDirectory + '/6_voipActivation2.png')
    #
    # voipIdentifyButton.click()
    #
    # ### Exede Voice - Part 1: Complete teh Voice Activation Portal Process Page
    # ### Step #2 - 911 Provisioning
    #
    # voipProvisionYesButton = WebDriverWait(driver, 60).until(
    #     EC.presence_of_element_located((By.XPATH, '//*[@href="/dap/3"]'))
    # )
    #
    # driver.save_screenshot(provioningScreenshotDirectory + '/7_voipActivation3.png')
    #
    # voipProvisionYesButton.click()
    # time.sleep(2)
    #
    # ### Exede Voice - Part 1: Complete teh Voice Activation Portal Process Page
    # ### Step #3 - Device
    #
    # deviceMacAddressField = WebDriverWait(driver, 60).until(
    #     EC.presence_of_element_located((By.ID, 'inputMacAddress'))
    # )
    #
    # alianzaMacAddress = "00A0BC4D9B52"
    # # this mac address is from the list provided by Alianza
    # deviceMacAddressField.send_keys(alianzaMacAddress)
    #
    # logger.debug("alianza Mac Address : " + alianzaMacAddress)
    #
    # time.sleep(1)
    #
    # deviceNextButton = WebDriverWait(driver, 60).until(
    #     EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/div[3]/div/div/div[2]/div/div[2]/div[2]/button'))
    # )
    #
    # driver.save_screenshot(provioningScreenshotDirectory + '/8_voipActivation4.png')
    #
    # deviceNextButton.click()
    #
    # ### Exede Voice - Part 1: Complete teh Voice Activation Portal Process Page
    # ### Step #4 - Summary
    #
    # ##//*[@class = "btn-default btn az-btn"][3]
    # ##//*[@id="root"]/div/div[3]/div/div/div[2]/div/div[2]/div[2]/button/span/text()
    # deviceActivateButton = WebDriverWait(driver, 60).until(
    #     EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Activate')]"))
    # )
    #
    # driver.save_screenshot(provioningScreenshotDirectory + '/9_voipActivation5.png')
    #
    # deviceActivateButton.click()
    #
    # logger.debug("activate button is clicked")
    #
    # time.sleep(2)
    # ### Exede Voice - Part 1: Complete the Voice Activation Portal Process Page
    # ### Exced Voice Activation Complete
    #
    # ExcedeVoiceButton = WebDriverWait(driver, 60).until(
    #     EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Exede Voice')]"))
    # )
    #
    # driver.save_screenshot(provioningScreenshotDirectory + '/10_voipActivation6.png')
    #
    # ExcedeVoiceButton.click()
    #
    # time.sleep(1)
    #
    # driver.switch_to_default_content()
    #
    # time.sleep(1)
    #
    # logger.debug("Back to the main IG window. driver.title : " + driver.title)
    #
    # logger.debug("after getting back to default content : " + driver.title)
    #
    # ###
    # verifyVoiceActivationButton = WebDriverWait(driver, 60).until(
    #     EC.presence_of_element_located((By.XPATH, '//input[@type="submit"]'))
    # )
    #
    # driver.save_screenshot(provioningScreenshotDirectory + '/11_voipActivation7.png')
    #
    # verifyVoiceActivationButton.click()

    time.sleep(6)

    ### Sometimes the next page is not displayed due to system error
    ### (Error Detail: Problem occurred updating the system with the provided information to register the modem. (#000048).
    try:
        customerButton = WebDriverWait(driver, 120).until(
            EC.element_to_be_clickable((By.XPATH, '//input[@value="Customer"]'))
        )
    except:
        logger.debug("Error Detail: Problem occurred updating the system with the provided information to register the modem. (#000048). Skipping to next row")
        continue

    customerButton.click()

    driver.save_screenshot(provioningScreenshotDirectory + '/5_newCustomerAccountSetup.png')

    lastFourField = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:paymentAuthentication"]'))
    )

    if paymentType.lower() == "CC".lower():
        lastFourField.send_keys("7777")
    elif paymentType.lower() == "EFT".lower():
        lastFourField.send_keys(bankAccountNumber[-4:])

    ccContinueButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//input[@value="Continue"]'))
    )

    ccContinueButton.click()

    time.sleep(5)

    # driver.switch_to_frame(1)

    pdfIFrame = driver.find_element_by_xpath('//*[@id="installerForm:j_id20"]/iframe')

    # logger.debug(pdfIFrame.get_attribute('src'))

    driver.switch_to_default_content()

    driver.switch_to_frame(pdfIFrame)

    time.sleep(5)

    driver.save_screenshot(provioningScreenshotDirectory + '/6_customerAgreement.png')

    time.sleep(2)

    getStartedButton = WebDriverWait(driver, 60).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="pnlElectronic"]/div/div[1]/button[1]/i'))
    )

    # logger.debug("getStartedButtonAttribute : " + getStartedButton.get_attribute('class'))

    getStartedButton.click()

    time.sleep(3)

    signField = driver.find_element_by_xpath('//*[@id="location1"]/div[2]/div[1]/input')

    logger.debug("signField type : " + signField.get_attribute('type'))

    signField.send_keys("Spider Man")

    time.sleep(3)

    driver.save_screenshot(provioningScreenshotDirectory + '/7_customerAgreementAfterSign.png')

    finishSubmitButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="completePopupContainer"]/div/div[1]/button'))
    )

    finishSubmitButton.click()

    time.sleep(2)

    driver.save_screenshot(provioningScreenshotDirectory + '/8_eSignSubmitted.png')

    time.sleep(2)

    driver.switch_to_default_content()

    continueButtonAfterSign = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id25"]'))
    )

    logger.debug('continueButtonAfterSign attribute : ' + continueButtonAfterSign.get_attribute('class'))

    driver.save_screenshot(provioningScreenshotDirectory + '/9_eSignComplete.png')

    continueButtonAfterSign.click()

    time.sleep(3)

    driver.save_screenshot(provioningScreenshotDirectory + '/10_activatingModem.png')

    # activatingModemContinueButton = WebDriverWait(driver, 120).until(
    #     EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id35"]'))
    # )

    activatingModemContinueButton = WebDriverWait(driver, 120).until(
        EC.presence_of_element_located((By.XPATH, '//input[@type="submit"]'))
    )

    activatingModemContinueButton.click()

    time.sleep(2)

    confirmationMessage = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id19"]'))
    )

    logger.debug("Confirmation Message After activation : " + confirmationMessage.text)

    driver.save_screenshot(provioningScreenshotDirectory + '/11_confirmation.png')

    # currentRow is moved to the next row
    currentRow = currentRow + 1

    ws.page_setup.fitToWidth = 1

    if currentRow < rowLength:
        logger.debug("Next row : " + str(currentRow-1))
        logger.debug('-----------------------------------------------------------------------')
    else:
        logger.debug("End of Run. All rows are processed.")

    assert "Success!" in driver.page_source

    try:
        wb.save('./Reports/NewConnectOrders.xlsx')
    except PermissionError:
        logger.debug('File is already open. Can\'t save')

    driver.quit()

logger.debug('-----------------------------------------------------------------------')
logger.debug('-----------------------------------------------------------------------')

# time.sleep(10)

# save to the project home directory
# driver.get_screenshot_as_file(".\\Screenshots\\facebook.png");

# logger.debug(driver.title)

# assert "Facebook" in driver.title

# driver.find_element_by_id("email").send_keys("Selenium Webdriver")
#
# driver.find_element_by_name("pass").send_keys("Python")
#
# driver.find_element_by_id("loginbutton").click()
#
# driver.quit()

def selectSalesChannel(x):
    return {
        'WB_DIRECT': 2,
        'ATT': 3,
        'MEXICO_RETAIL': 4,
        'US_COMMUNITIES': 5,
        'US_SMALLBUSINESS':6,
        'DISH_DIRECT_RETAIL':7,
        'MEDIA_NETWORKS':8,
        'B2B_PARTNERS':9,
        'FIELD_TRIAL':10
    }.get(x, 'WB_DIRECT')