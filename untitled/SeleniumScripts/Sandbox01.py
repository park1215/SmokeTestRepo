__author__='SeanPark_ViaSat'

import datetime
import time
import string
import random
import openpyxl
import pandas as pd
import xlsxwriter
import xlwt
import os

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import selenium.webdriver.support.ui as ui
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys

bankRoutingNumber = '122105278'

bankAccountNumber = '0000000016'

paymentType = ''

wb = openpyxl.load_workbook('./Reports/NewConnectOrders.xlsx')

wbAddress = openpyxl.load_workbook('./Data/Addresses.xlsx')

sheetAddress = wbAddress['Sheet1']

rowLength = sheetAddress.max_row

columnLength = sheetAddress.max_column

currentRow = 2

for item in range(0, 1):
    print("Number of Orders to be created : " + str(rowLength-1))
    print("currentRow : " + str(currentRow-1))

    username = sheetAddress.cell(row=currentRow, column=1).value

    password = sheetAddress.cell(row=currentRow, column=2).value

    salesChannel = sheetAddress.cell(row=currentRow, column=3).value

    customerType = sheetAddress.cell(row=currentRow, column=4).value

    addressLine1 = sheetAddress.cell(row=currentRow, column=7).value

    city = sheetAddress.cell(row=currentRow, column=9).value

    state = sheetAddress.cell(row=currentRow, column=10).value

    zipCode = sheetAddress.cell(row=currentRow, column=11).value

    paymentType = sheetAddress.cell(row=currentRow, column=18).value

    packageName = sheetAddress.cell(row=currentRow, column=25).value

    satelliteName = sheetAddress.cell(row=currentRow, column=26).value

    print("Payment type : " + paymentType)

    print("Package Name : " + packageName)

    print("Satellite Name : " + satelliteName)

    driver = webdriver.Chrome("C:\\Selenium\\chromedriver.exe")
    # driver = webdriver.Ie("C:\Selenium\IEDriverServer.exe");

    driver.implicitly_wait(50)

    driver.set_page_load_timeout(30)

    driver.get("https://ordermgmt.test.exede.net/PublicGUI-SupportGUI/v1/pages/addcustomer/serviceAvailability.xhtml")

    driver.maximize_window()

    driver.implicitly_wait(20)

    driver.find_element_by_xpath(
        "//*[@id=\"document:body\"]/table/tbody/tr[2]/td/form/table/tbody/tr[3]/td[2]/input").send_keys(username)

    driver.find_element_by_xpath(
        "//*[@id=\"document:body\"]/table/tbody/tr[2]/td/form/table/tbody/tr[4]/td[2]/input").send_keys(password)

    driver.find_element_by_name("submit").click()

    driver.implicitly_wait(5)

    # addCustomerTab = WebDriverWait(driver, 10).until(
    #     EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:add"]'))
    # )

    addCustomerTab = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="add"]'))
    )

    addCustomerTab.click()

    driver.implicitly_wait(3)
    time.sleep(1)

    # if salesChannel == 'WB_DIRECT':
    #     driver.find_element_by_xpath("//*[@id=\"addCustomerForm:salesChannelMenu\"]/option[2]").click()

    if salesChannel == 'WB_DIRECT':
        # driver.find_element_by_xpath("//*[@id=\"addCustomerForm:salesChannelMenu\"]/option[2]").click()
        salesChannelSelected = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:salesChannelMenu"]/option[2]'))
        )
    salesChannelSelected.click()

    if customerType.lower() == 'Residential'.lower():
        # driver.find_element_by_xpath("//*[@id=\"addCustomerForm:salesChannelMenu\"]/option[2]").click()
        customerTypeSelected = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:customerTypeMenu"]/option[1]'))
        )

    customerTypeSelected.click()

    # salesChannelOption = selectSalesChannel(salesChannel)
    # select = Select(driver.find_element_by_id('addCustomerForm:salesChannelMenu'))
    #
    # select.select_by_visible_text(salesChannel).click()

    now = datetime.datetime.now()

    currentYear = str(now.year)

    months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']

    hexdigits = list(string.hexdigits)
    del hexdigits[10:16]

    # print(hexdigits)

    randomMac = "AA:BB:CC:"

    for x in range(0, 6):
        randomNumber = random.choice(hexdigits)
        randomMac = randomMac + randomNumber
        if x % 2 != 0 and len(randomMac) < 17:
            randomMac = randomMac + ":"

    print("Mac Address : " + randomMac)

    randomMacNoColon = randomMac.replace(':', '')

    print(randomMacNoColon)

    currentMonth = months[now.month - 1]

    currentDay = ""

    if now.day < 10:
        currentDay = '0' + str(now.day)
    else:
        currentDay = str(now.day)

    # transactionReference = "SPark_" + currentDay + currentMonth + currentYear + str(item+15)

    transactionReference = "SPark_" + str(randomMacNoColon)

    newSheetName = currentDay + "-" + currentMonth + "-" + currentYear

    sheetList = wb.sheetnames

    if newSheetName not in sheetList:
        print(newSheetName + " Not exist")
        wb.create_sheet(newSheetName)

    ws = wb[newSheetName]

    end_of_sheet = ws.max_row

    start_of_sheet = end_of_sheet + 1

    print("Start of sheet " + str(start_of_sheet))

    ws.cell(row=1, column=2).value = 'Transaction Reference'
    ws.cell(row=1, column=3).value = 'Service Agreement'
    ws.cell(row=1, column=4).value = 'MAC'
    ws.cell(row=1, column=5).value = 'Sales Channel'
    ws.cell(row=1, column=6).value = 'Customer Type'
    ws.cell(row=1, column=7).value = 'Payment Type'
    ws.cell(row=1, column=8).value = 'Satellite'
    ws.cell(row=1, column=9).value = 'Package'
    ws.cell(row=1, column=10).value = 'Device'
    ws.cell(row=1, column=11).value = 'Voip'

    ws.cell(row=start_of_sheet, column=2).value = transactionReference
    ws.cell(row=start_of_sheet, column=4).value = randomMac
    ws.cell(row=start_of_sheet, column=5).value = salesChannelSelected.text
    ws.cell(row=start_of_sheet, column=6).value = customerTypeSelected.text

    try:
        wb.save('./Reports/NewConnectOrders.xlsx')
    except PermissionError:
        print('File is already open. Can\'t save')

    print("end of process")

'''
    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:transactionReference\"]").send_keys(transactionReference)

    # driver.implicitly_wait(2)
    # driver.find_element_by_xpath("//*[@id=\"addCustomerForm:namesIdName1\"]").send_keys("Spider")

    time.sleep(1)
    firstNameField = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID,
                                        'addCustomerForm:namesIdName1'))
    )

    firstNameField.send_keys("Spider")

    driver.implicitly_wait(2)

    time.sleep(1)
    lastNameField = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID,
                                        'addCustomerForm:namesIdName3'))
    )

    lastNameField.send_keys("Man")

    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableAddress1\"]").send_keys(addressLine1)
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableCity\"]").send_keys(city)
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableStateAddressState\"]/option[7]").click()
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableZip\"]").send_keys(zipCode)
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:primaryPhoneIdMaybeTablePhoneNumber\"]").send_keys("7204823823")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:emailAddressId\"]").send_keys("sean.park@viasat.com")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:Birthdate\"]").send_keys("01/01/1980")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:nextButtonId\"]").click()
    time.sleep(1)

    # Contacts Page

    # creditCheckPassed = WebDriverWait(driver, 60).until(
    # EC.presence_of_element_located((By.XPATH, '// *[@id = "addCustomerForm:_id91"]/tbody/tr/td/span'))
    # )

    driver.implicitly_wait(2)

    # assert "The Credit Check passed." in driver.page_source

    time.sleep(1)

    customerReferenceField = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID,
                                        'addCustomerForm:customerReference'))
    )

    customerReferenceField.send_keys(transactionReference)

    driver.implicitly_wait(2)
    time.sleep(1)
    accountReferenceField = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID,
                                        'addCustomerForm:accountReference'))
    )

    accountReferenceField.send_keys(transactionReference)

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:nextButtonId\"]").click()

    # Packages Page

    packagesTitle = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:packagesHeaderLabel"]'))
    )

    # Sometimes, radio button is not checked by default. So intentionally click it.
    # id is dynamically created. id114 or id112
    # driver.find_element_by_xpath('//*[@id="addCustomerForm:_id114:_2"]').click()
    # driver.find_element_by_xpath('//*[@id="addCustomerForm:_id112:_2"]').click()

    time.sleep(1)

    # This statement is to be replaced with dictionary type switch statement.
    # if packageName.lower() == 'Unlimited Bronze 12'.lower():
    #     packageRadioButton = WebDriverWait(driver, 20).until(
    #         EC.presence_of_element_located((By.XPATH, '//input[starts-with(@value, "$")]'))
    #     )

    if packageName.lower() == 'Unlimited Bronze 12'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_0'))
        )

    elif packageName.lower() == 'Unlimited Silver 12'.lower() and satelliteName.lower() == 'VS1'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_1"]'))
        )

    elif packageName.lower() == 'Unlimited Gold 12'.lower() and satelliteName.lower() == 'VS1'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_2"]'))
        )

    elif packageName.lower() == 'Unlimited Silver 25'.lower() and satelliteName.lower() == 'VS1'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_4"]'))
        )

    elif packageName.lower() == 'Unlimited Gold 30'.lower() and satelliteName.lower() == 'VS1'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_5"]'))
        )

    elif packageName.lower() == 'Unlimited Silver 25'.lower() and satelliteName.lower() == 'VS2'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_1"]'))
        )

    elif packageName.lower() == 'Unlimited Gold 50'.lower() and satelliteName.lower() == 'VS2'.lower():
        packageRadioButton = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:topPackages:_2"]'))
        )

    packageRadioButton.click()

    print("process is ended")
    '''

