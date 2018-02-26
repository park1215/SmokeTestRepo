__author__ = 'Sean Park_ViaSat'

__author__='SeanPark_ViaSat'

import datetime
import time
import string
import random
import openpyxl
import pandas as pd
import xlsxwriter
import xlwt
import os

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import selenium.webdriver.support.ui as ui
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys

'''Creating a new file, using openpyxl
from openpyxl import Workbook
import time

book = Workbook()
sheet = book.active

sheet['A1'] = 56
sheet[A2] = 43

now = time.stringftime("%x") #02/02/2018
sheet['A3'] = now

book.save("sample.xlsx")
'''

'''Reading a file, using openpyxl
book = openpyxl.load_workbook('sample.xlsx')

sheet = book.active

a1 = sheet['A1']
a2 = sheet['A2']
a3 = sheet.cell(row=3, column=1)

# Reading multiple cells
cells = sheet['A1': 'B6']
for c1, c2 in cells:
print("{0:8} {1:8}".format(c1.value, c2.value))

#iterating_by_rows
rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)
    
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print() 
    
#Sheets
import openpyxl
book = openpyxl.load_workbook('sheets.xlsx')

print(book.get_sheet_names())

active_sheet = book.active
print(type(ative_sheet))

sheet - book.get_sheet_by_name("March")
print(sheet.title)
'''
wb = openpyxl.load_workbook('.\Data\Addresses.xlsx')

print(os.getcwd())

print(type(wb))

print(wb.sheetnames)

# wb.create_sheet('1')
sheet = wb['Sheet1']

# sheet.cell(row=10, column=2).value = 'test'

print("max row : " + str(sheet.max_row))
print("max column : " + str(sheet.max_column))

# sheet.title = 'superman'

print(wb.sheetnames)

print(type(sheet))

print(sheet.cell(row=2, column=2).value)

sheetAddress = wb['Sheet1']

print(sheetAddress.cell(row=4, column=2).value)
print(sheetAddress.cell(row=4, column=3).value)
print(sheetAddress.cell(row=4, column=4).value)
print(sheetAddress.cell(row=4, column=5).value)

rowLength = sheetAddress.max_row

columnLength = sheetAddress.max_column

currentRow = 2

for item in range(currentRow, rowLength):

    username = sheetAddress.cell(row=currentRow, column=1).value

    password = sheetAddress.cell(row=currentRow, column=2).value

    salesChannel = sheetAddress.cell(row=currentRow, column=3).value

    customerType = sheetAddress.cell(row=currentRow, column=4).value

    addressLine1 = sheetAddress.cell(row=currentRow, column=7).value

    city = sheetAddress.cell(row=currentRow, column=9).value

    state = sheetAddress.cell(row=currentRow, column=10).value

    zipCode = sheetAddress.cell(row=currentRow, column=11).value

    print("username : " + username)

    print("password : " + password)

    print("salesChannel : " + salesChannel)

    print("customerType : " + customerType)

    print("addressLine1 : " + addressLine1)

    print("city : " + city)

    print("state : " + state)

    print("zipCode : " + str(zipCode))

    currentRow = currentRow + 1

#sheet.cell(row=0, column=0).value = 2
# test2sheet['A1'] = 10

# book.create_sheet("test2", 2)
#
# book.create_sheet("test3")



#book.create_sheet("Test1", 2)

# book.save("NewConnectOrders.xlsx")

# # Excel writing/reading for the result of the submitted orders.
# wb = Workbook()
# sheet1 = wb.add_sheet('Sheet1')
# sheet2 = wb.add_sheet('Sheet2')
# sheet3 = wb.add_sheet('Sheet3')
#
# # sheet1.write(0, 0, 'This is Sheet1')
# # sheet2.write(0, 0, 'This is Sheet2')
# # sheet3.write(0, 0, 'This is Sheet3')
#
# # Change cell width
# sheet1.col(0).width = 7000
# sheet2.col(0).width = 7000
# sheet3.col(0).width = 7000
#
# # Coloring Cells
# style1 = easyxf('pattern: pattern solid, fore_colour yellow;')
#
# # Inserting formulas
# for x in range(0, 10):
#     sheet1.write(x, 0, x)
#
# sheet1.write(10, 0, Formula('SUM(A1:A10)'), style1)
#
# wb.save('xlwt_example.xlsx')

# xlrd
# import xlrd
#
# workbook = xlrd.open_workbook("UsersBook.xlsx")
#
# worksheet = workbook.sheet_by_name("Users")
# worksheet = workbook.sheet_by_index(0)
#
# print("the value at row 4 and column 2 is : {0}".format(worksheet.cell(4, 2).value))
#
# sheet_count = workbook.nsheets
# print("the total number of sheets : {0}".format(sheet_count))
#
# sheet_names = workbook.sheet_names()
# print("sheet names : {0}".format(sheets_name))
# sheet names : ['Users', 'Nothing Here']

#to find the total number of rows and columns in the sheet, user the property nrows and ncols with
#the sheet object

# total_rows = worksheet.nrows
# total_cols = worksheet.ncols
# print("number of rows: {}, and number of columns : {1}".format(total_rows, total_cols))
#
# # loop in every record in the worksheet and store them in a list then display the final list:
# table = list()
# record = list()
#
# for x in range(total_rows):
#     for y in range(total_cols):
#         record.append(worksheet.cell(x, y).value)
#     table.append(record)
#     record = []
#     x += 1
#
# print(table)

# from xutils module, call copy class
# from xutils.copy import copy

# open the excel file
# rb = xlrd.open_workbook('UsersBook.xls')

# make a writable copy of the opened excel file
# wb = copy(rb)

# read the first sheet to write to within the writable copy
# w_sheet = wb.get_sheet(0)

# write or modify the value at first row = second column
# w_sheet.write(0, 1, 'Modified!')

# the last step saving the workbook
# wb.save('UsersBook.xls')

# Create a Pandas dataframe from the data
# df = pd.DataFrame({'Sales Channel':['test']})

# Create a Pandas Excel writer using XlsxWriter as the engine.
# writer = pd.ExcelWriter('Pandas_simple.xlsx', engine='xlsxwriter')

# Convert the dataframe to an XlsxWriter Excel object
# df.to_excel(writer, sheet_name='Template')

# Close the Pandas Excel writer and output the Excel file
# writer.save()

