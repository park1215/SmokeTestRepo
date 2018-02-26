__author__='SeanPark_ViaSat'

import datetime
import time
import string
import random
import openpyxl
import pandas as pd
import xlsxwriter
import xlwt
import os as winos

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys

wbAddress = openpyxl.load_workbook('./Data/Addresses.xlsx')

wb = openpyxl.load_workbook('./Reports/NewConnectOrders.xlsx')

sheetAddress = wbAddress['Sheet1']

rowLength = sheetAddress.max_row

columnLength = sheetAddress.max_column

currentRow = 2

for item in range(currentRow, rowLength+1):
    print("Number of Orders to be created in this test: " + str(rowLength-1))

    print("Current Row in address list: " + str(currentRow-1))

    username = sheetAddress.cell(row=currentRow, column=1).value

    password = sheetAddress.cell(row=currentRow, column=2).value

    salesChannel = sheetAddress.cell(row=currentRow, column=3).value

    customerType = sheetAddress.cell(row=currentRow, column=4).value

    addressLine1 = sheetAddress.cell(row=currentRow, column=7).value

    city = sheetAddress.cell(row=currentRow, column=9).value

    state = sheetAddress.cell(row=currentRow, column=10).value

    zipCode = sheetAddress.cell(row=currentRow, column=11).value

    voipIncluded = sheetAddress.cell(row=currentRow, column=24).value

    if voipIncluded.lower() == "no":
        voipIncluded = False

    elif voipIncluded.lower() == "yes":
        voipIncluded = True
    else:
        voipIncluded = None

    print("void Include? : " + str(voipIncluded))

    driver = webdriver.Chrome("C:\\Selenium\\chromedriver.exe")
    # driver = webdriver.Ie("C:\Selenium\IEDriverServer.exe");

    driver.implicitly_wait(50)

    driver.set_page_load_timeout(30)

    driver.get("https://ordermgmt.test.exede.net/PublicGUI-SupportGUI/v1/pages/addcustomer/serviceAvailability.xhtml")

    driver.maximize_window()

    driver.implicitly_wait(20)

    driver.find_element_by_xpath(
        "//*[@id=\"document:body\"]/table/tbody/tr[2]/td/form/table/tbody/tr[3]/td[2]/input").send_keys(username)

    driver.find_element_by_xpath(
        "//*[@id=\"document:body\"]/table/tbody/tr[2]/td/form/table/tbody/tr[4]/td[2]/input").send_keys(password)

    driver.find_element_by_name("submit").click()

    driver.implicitly_wait(5)

    # addCustomerTab = WebDriverWait(driver, 10).until(
    #     EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:add"]'))
    # )

    addCustomerTab = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="add"]'))
    )

    addCustomerTab.click()

    driver.implicitly_wait(3)
    time.sleep(1)

    if salesChannel == 'WB_DIRECT':
        driver.find_element_by_xpath("//*[@id=\"addCustomerForm:salesChannelMenu\"]/option[2]").click()

    # salesChannelOption = selectSalesChannel(salesChannel)
    # select = Select(driver.find_element_by_id('addCustomerForm:salesChannelMenu'))
    #
    # select.select_by_visible_text(salesChannel).click()

    now = datetime.datetime.now()

    currentYear = str(now.year)

    months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']

    hexdigits = list(string.hexdigits)
    del hexdigits[10:16]

    # print(hexdigits)

    randomMac = "AA:BB:CC:"

    for x in range(0, 6):
        randomNumber = random.choice(hexdigits)
        randomMac = randomMac + randomNumber
        if x % 2 != 0 and len(randomMac) < 17:
            randomMac = randomMac + ":"

    print("Mac Address : " + randomMac)

    randomMacNoColon = randomMac.replace(':', '')

    print(randomMacNoColon)

    currentMonth = months[now.month - 1]

    currentDay = ""

    if now.day < 10:
        currentDay = '0' + str(now.day)
    else:
        currentDay = str(now.day)

    # transactionReference = "SPark_" + currentDay + currentMonth + currentYear + str(item+15)

    transactionReference = "SPark_" + str(randomMacNoColon)

    SupportPortalScreenshotDirectory = './Reports/' + randomMacNoColon

    if not winos.path.exists(SupportPortalScreenshotDirectory):
        winos.makedirs(SupportPortalScreenshotDirectory)

    newSheet = currentDay + "-" + currentMonth + "-" + currentYear

    wb.create_sheet(newSheet)

    ws = wb[newSheet]

    ws.cell(row=1, column=2).value = 'Transaction Reference'
    ws.cell(row=1, column=3).value = 'Service Agreement'
    ws.cell(row=1, column=4).value = 'MAC'
    ws.cell(row=2, column=2).value = transactionReference
    ws.cell(row=2, column=4).value = randomMac

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:transactionReference\"]").send_keys(transactionReference)

    # driver.implicitly_wait(2)
    # driver.find_element_by_xpath("//*[@id=\"addCustomerForm:namesIdName1\"]").send_keys("Spider")

    time.sleep(1)
    firstNameField = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID,
                                        'addCustomerForm:namesIdName1'))
    )

    firstNameField.send_keys("Spider")

    driver.implicitly_wait(2)

    time.sleep(1)
    lastNameField = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID,
                                        'addCustomerForm:namesIdName3'))
    )

    lastNameField.send_keys("Man")

    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableAddress1\"]").send_keys(addressLine1)
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableCity\"]").send_keys(city)
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableStateAddressState\"]/option[7]").click()
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableZip\"]").send_keys(zipCode)
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:primaryPhoneIdMaybeTablePhoneNumber\"]").send_keys("7204823823")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:emailAddressId\"]").send_keys("sean.park@viasat.com")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:Birthdate\"]").send_keys("12/15/1973")
    time.sleep(1)

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/1_serviceability.png')

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:nextButtonId\"]").click()
    time.sleep(1)

    # Contacts Page

    # creditCheckPassed = WebDriverWait(driver, 60).until(
    # EC.presence_of_element_located((By.XPATH, '// *[@id = "addCustomerForm:_id91"]/tbody/tr/td/span'))
    # )

    driver.implicitly_wait(2)

    # assert "The Credit Check passed." in driver.page_source

    time.sleep(1)

    customerReferenceField = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID,
                                        'addCustomerForm:customerReference'))
    )

    customerReferenceField.send_keys(transactionReference)

    driver.implicitly_wait(2)
    time.sleep(1)
    accountReferenceField = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID,
                                        'addCustomerForm:accountReference'))
    )

    accountReferenceField.send_keys(transactionReference)

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/2_contacts.png')

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:nextButtonId\"]").click()

    # Packages Page

    packagesTitle = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:packagesHeaderLabel"]'))
    )

    # Sometimes, radio button is not checked by default. So intentionally click it.
    try:
        driver.find_element_by_xpath('//*[@id="addCustomerForm:_id112:_2"]').click()
    except:
        print("radioButtonID is not id112")

    try:
        driver.find_element_by_xpath('//*[@id="addCustomerForm:_id114:_2"]').click()
    except:
        print("radioButtonID is not id114")

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/3_packages.png')

    # "is not clickable at point" error. Another element is covering the element to click. I could use execute_script() to click on this.
    nextButton = driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]')
    driver.execute_script("arguments[0].click();", nextButton)

    driver.implicitly_wait(2)
    time.sleep(1)

    # Options Page
    optionsTitle = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:optionsLabel"]'))
    )

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:_1selectionPackages:_1\"]").click()

    driver.implicitly_wait(2)
    time.sleep(1)

    voipSelectButton = WebDriverWait(driver, 180).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:_id248"]'))
    )

    voipSelectButton.click()

    driver.implicitly_wait(2)
    time.sleep(1)
    # driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]').click()

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/4_options.png')

    optionsPageNextButton = driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]')
    driver.execute_script("arguments[0].click();", optionsPageNextButton)

    driver.implicitly_wait(2)
    time.sleep(1)

    #VoIP Page

    driver.implicitly_wait(2)
    time.sleep(1)

    voipPassword = "Qkrtmd!1"

    voipPasswordField = WebDriverWait(driver, 180).until(
        EC.presence_of_element_located((By.ID, 'addCustomerForm:password'))
    )

    voipPasswordField.send_keys(voipPassword)

    driver.implicitly_wait(2)
    time.sleep(1)

    print("voipUserName : " + driver.find_element_by_xpath('//*[@id="addCustomerForm:userName"]').text)

    voipPasswordConfirmField = WebDriverWait(driver, 180).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:confirm"]'))
    )

    voipPasswordConfirmField.send_keys(voipPassword)

    driver.implicitly_wait(2)
    time.sleep(1)

    streetAddressNumberField = WebDriverWait(driver, 180).until(
        EC.presence_of_element_located((By.ID, 'addCustomerForm:streetNumber'))
    )

    streetAddressNumberField.send_keys("12017")

    driver.implicitly_wait(2)
    time.sleep(1)

    streetAddressStreetNameField = WebDriverWait(driver, 180).until(
        EC.presence_of_element_located((By.ID, 'addCustomerForm:streetName'))
    )

    streetAddressStreetNameField.send_keys("E Lake Cir")

    driver.implicitly_wait(2)
    time.sleep(1)

    phoneNumberSelected = WebDriverWait(driver, 180).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:voipChooseNumberMenu"]/option[2]'))
    )

    phoneNumberSelected.click()

    voipNextButton = driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]')

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/5_voip.png')

    driver.execute_script("arguments[0].click();", voipNextButton)

    # Payment Page
    driver.implicitly_wait(2)
    time.sleep(5)

    try:
        paymentMethodTitle = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:recurringPaymentInfoLabel"]'))
        )
    except:
        print("")

    driver.implicitly_wait(2)
    time.sleep(3)

    creditCardType = WebDriverWait(driver, 180).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id=\"addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdcreditCardTypeId\"]/option[3]"))
    )

    creditCardType.click()

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdNumberId\"]").send_keys("4012000077777777")

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdExpireMonthIdMonthId\"]/option[5]").click()

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdExpireYearIdYearId"]/option[3]').click()

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdFirstNameId"]').send_keys("VISA")

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdLastNameId"]').send_keys("APPROVAL")

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdAddressZip"]').send_keys("80111")

    try:
        taxJurisdictionDropdown = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:taxJurisdictionMenu"]'))
        )

        driver.find_element_by_xpath('//*[@id="addCustomerForm:taxJurisdictionMenu"]/option[2]').click()

        print("taxJurisdiction is a dropdown menu")
    except:
        print("taxJurisdiction is NOT a dropdown menu")

    # dropdownPresent = driver.find_element_by_xpath('//*[@id="addCustomerForm:taxJurisdictionMenu"]/option[2]').is_displayed()

    # if dropdownPresent:
    #     driver.find_element_by_xpath('//*[@id="addCustomerForm:taxJurisdictionMenu"]/option[2]').click()

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/5_payment.png')

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]').click()

    # Review Page

    scheduleButton = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:scheduleInstallationButtonId"]'))
    )

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/6_review.png')

    scheduleButton.click()
    time.sleep(2)

    # Schedule Page

    submitOrderButton = WebDriverWait(driver, 50).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:submitButtonId"]'))
    )

    time.sleep(1)

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/7_schedule.png')

    submitOrderButton.click()


    # wait for order reference number created

    # Confirmation Page

    printButton = WebDriverWait(driver, 180).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:printButtonId"]'))
    )

    serviceAgreementReference = driver.find_element_by_xpath('//*[@id="addCustomerForm:serviceAgreementReference"]').text

    print("Sales Channel : " + salesChannel)
    print("External Account Reference : " + serviceAgreementReference)

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/8_confirmation.png')

    newOrderButton = driver.find_element_by_xpath('//*[@id="addCustomerForm:newOrderButtonId"]')

    newOrderButton.click()

    driver.implicitly_wait(20)
    time.sleep(1)

    transactionInfoTitle = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:transactionInfoLabel"]'))
    )

    driver.get('https://spyglass01.test.wdc1.wildblue.net:8443/SpyGlass/')

    referenceType = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[1]/select/option[5]'))
        )

    referenceType.click()

    referenceValue = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[2]/input'))
        )

    referenceValue.send_keys(transactionReference)

    externalSystem = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[3]/div/select/option[2]'))
        )

    externalSystem.click()

    driver.implicitly_wait(20)
    time.sleep(1)

    driver.find_element_by_xpath('/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[4]/input[1]').click()

    fsmCustomerCode = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="datatable"]/tbody/tr[1]/td[1]/div[1]'))
        )

    serviceAgreementNumber = driver.find_element_by_xpath('//*[@id="data"]/table[1]/tbody/tr[2]/td/table/tbody/tr[2]/td[12]').text

    print('serviceAgreementNumber : ' + serviceAgreementNumber)

    driver.save_screenshot(SupportPortalScreenshotDirectory + '/9_spyglass.png')

    driver.implicitly_wait(20)
    time.sleep(1)

    ws.cell(row=2, column=3).value = serviceAgreementNumber

    #######################################
    ###
    ### Provisioning Starting from here....
    ###
    #######################################

    # driver = webdriver.Chrome("C:\\Selenium\\chromedriver.exe")
    # driver = webdriver.Ie("C:\\Selenium\\IEDriverServer.exe")

    # driver.set_page_load_timeout(30)

    time.sleep(6)

    # driver.get("https://ordermgmt.test.exede.net/PublicGUI-SupportGUI/v1/pages/addcustomer/serviceAvailability.xhtml")
    #
    # driver.get('https://spyglass01.test.wdc1.wildblue.net:8443/SpyGlass/')

    installGUI = "https://igui-installationgui.test.wdc1.wildblue.net/InternalGUI-InstallationGUI/"

    installGUIwithMac = installGUI + "?n=" + randomMacNoColon

    driver.get(installGUIwithMac)

    provioningScreenshotDirectory = SupportPortalScreenshotDirectory + '/' + serviceAgreementNumber

    if not winos.path.exists(provioningScreenshotDirectory):
        winos.makedirs(provioningScreenshotDirectory)

    driver.maximize_window()

    time.sleep(3)

    print("Web Browser in test : " + driver.name)

    ### if it's IE, it needs to bypass security warning
    if driver.name == "internet explorer":
        print("driver is IE")
        continueLink = driver.find_element_by_id('overridelink')
        continueLink.click()

    driver.implicitly_wait(20)

    activationKeyField = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:activationKey"]'))
    )

    activationKeyField.send_keys(serviceAgreementNumber)

    installButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id36"]'))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/1_welcomeToServiceActivation.png')

    installButton.click()

    time.sleep(2)

    ### Customer confirmation New Installation Page

    installerNumberField = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:installerId"]'))
    )

    installerNumberField.send_keys("99072761")

    driver.save_screenshot(provioningScreenshotDirectory + '/2_customerConfirmationNewInstallation.png')

    continueInstallButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id53"]'))
    )

    continueInstallButton.click()

    time.sleep(5)

    ### Email confirmation & Update Page
    emailConfirmationButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id30"]'))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/3_emailConfirmationAndUpdate.png')

    emailConfirmationButton.click()

    time.sleep(10)

    ### Quality of Install Page
    # qOIcontinueButton = WebDriverWait(driver, 60).until(
    #     EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id50"]'))
    # )

    thankYouTag = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id40"]'))
    )

    print('Thank You Tag is displayed : ' + thankYouTag.text)

    qOIcontinueButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//input[@type="submit"]'))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/4_qualityOfInstall.png')

    qOIcontinueButton.click()

    ### Exede Voice Page
    print('Entering Exede Voice if voip added to service...')

    voiceActivationPortalButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@value="Voice Activation Portal"]'))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/5_voipActivation1.png')

    voiceActivationPortalButton.click()

    ### Exede Voice - Part 1: Complete teh Voice Activation Portal Process Page
    ### Step #1 - Identify Account Page
    time.sleep(2)
    voipIFrame = driver.find_element_by_xpath('//*[@id="installerForm:j_id25"]/iframe')

    time.sleep(2)

    driver.switch_to_default_content()

    print("default content iFrame driver title : " + driver.title)

    driver.switch_to_frame(voipIFrame)

    print("VoiP activation iFrame entered...")

    print("VoiP iFrame driver title : " + driver.title)

    voipAccountNumberField = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.ID, 'inputAccountNumber'))
    )

    voipAccountNumberField.send_keys(serviceAgreementNumber)

    voipLastNameField = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.ID, 'inputLastName'))
    )

    voipLastNameField.send_keys("Man")

    voipIdentifyButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@type="submit"]'))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/6_voipActivation2.png')

    voipIdentifyButton.click()

    ### Exede Voice - Part 1: Complete teh Voice Activation Portal Process Page
    ### Step #2 - 911 Provisioning

    voipProvisionYesButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@href="/dap/3"]'))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/7_voipActivation3.png')

    voipProvisionYesButton.click()
    time.sleep(2)

    ### Exede Voice - Part 1: Complete teh Voice Activation Portal Process Page
    ### Step #3 - Device

    deviceMacAddressField = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.ID, 'inputMacAddress'))
    )

    alianzaMacAddress = "00A0BC4D9B5E"
    # this mac address is from the list provided by Alianza
    deviceMacAddressField.send_keys(alianzaMacAddress)

    print("alianza Mac Address : " + alianzaMacAddress)

    time.sleep(1)

    deviceNextButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/div[3]/div/div/div[2]/div/div[2]/div[2]/button'))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/8_voipActivation4.png')

    deviceNextButton.click()

    ### Exede Voice - Part 1: Complete teh Voice Activation Portal Process Page
    ### Step #4 - Summary

    ##//*[@class = "btn-default btn az-btn"][3]
    ##//*[@id="root"]/div/div[3]/div/div/div[2]/div/div[2]/div[2]/button/span/text()
    deviceActivateButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'Activate')]"))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/9_voipActivation5.png')

    deviceActivateButton.click()

    print("activate button is clicked")

    time.sleep(2)
    ### Exede Voice - Part 1: Complete the Voice Activation Portal Process Page
    ### Exced Voice Activation Complete

    ExcedeVoiceButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Exede Voice')]"))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/10_voipActivation6.png')

    ExcedeVoiceButton.click()

    time.sleep(1)

    driver.switch_to_default_content()

    time.sleep(1)

    print("Back to the main IG window. driver.title : " + driver.title)

    print("after getting back to default content : " + driver.title)

    ###
    verifyVoiceActivationButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//input[@type="submit"]'))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/11_voipActivation7.png')

    verifyVoiceActivationButton.click()

    time.sleep(6)

    customerButton = WebDriverWait(driver, 60).until(
        EC.element_to_be_clickable((By.XPATH, '//input[@value="Customer"]'))
    )

    customerButton.click()

    driver.save_screenshot(provioningScreenshotDirectory + '/12_newCustomerAccountSetup.png')

    lastFourField = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:paymentAuthentication"]'))
    )

    lastFourField.send_keys("7777")

    ccContinueButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//input[@value="Continue"]'))
    )

    ccContinueButton.click()

    time.sleep(5)

    # driver.switch_to_frame(1)

    pdfIFrame = driver.find_element_by_xpath('//*[@id="installerForm:j_id20"]/iframe')

    # print(pdfIFrame.get_attribute('src'))

    driver.switch_to_default_content()

    driver.switch_to_frame(pdfIFrame)

    time.sleep(5)

    driver.save_screenshot(provioningScreenshotDirectory + '/13_customerAgreement.png')

    time.sleep(2)

    getStartedButton = WebDriverWait(driver, 60).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="pnlElectronic"]/div/div[1]/button[1]/i'))
    )

    print("getStartedButtonAttribute : " + getStartedButton.get_attribute('class'))

    getStartedButton.click()

    time.sleep(3)

    signField = driver.find_element_by_xpath('//*[@id="location1"]/div[2]/div[1]/input')

    print(signField.get_attribute('type'))

    signField.send_keys("Spider Man")

    time.sleep(3)

    driver.save_screenshot(provioningScreenshotDirectory + '/14_customerAgreementAfterSign.png')

    finishSubmitButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="completePopupContainer"]/div/div[1]/button'))
    )

    finishSubmitButton.click()

    time.sleep(2)

    driver.save_screenshot(provioningScreenshotDirectory + '/15_eSignSubmitted.png')

    time.sleep(2)

    driver.switch_to_default_content()

    continueButtonAfterSign = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id25"]'))
    )

    print('continueButtonAfterSign attribute : ' + continueButtonAfterSign.get_attribute('class'))

    driver.save_screenshot(provioningScreenshotDirectory + '/16_eSignComplete.png')

    continueButtonAfterSign.click()

    time.sleep(3)

    driver.save_screenshot(provioningScreenshotDirectory + '/17_activatingModem.png')

    activatingModemContinueButton = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id35"]'))
    )

    activatingModemContinueButton.click()

    time.sleep(2)

    confirmationMessage = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="installerForm:j_id19"]'))
    )

    driver.save_screenshot(provioningScreenshotDirectory + '/18_confirmation.png')

    assert "Success!" in driver.page_source

    driver.quit()

wb.save('NewConnectOrders.xlsx')





# time.sleep(10)

# save to the project home directory
# driver.get_screenshot_as_file(".\\Screenshots\\facebook.png");

# print(driver.title)

# assert "Facebook" in driver.title

# driver.find_element_by_id("email").send_keys("Selenium Webdriver")
#
# driver.find_element_by_name("pass").send_keys("Python")
#
# driver.find_element_by_id("loginbutton").click()
#
# driver.quit()

def selectSalesChannel(x):
    return {
        'WB_DIRECT': 2,
        'ATT': 3,
        'MEXICO_RETAIL': 4,
        'US_COMMUNITIES': 5,
        'US_SMALLBUSINESS':6,
        'DISH_DIRECT_RETAIL':7,
        'MEDIA_NETWORKS':8,
        'B2B_PARTNERS':9,
        'FIELD_TRIAL':10
    }.get(x, 'WB_DIRECT')