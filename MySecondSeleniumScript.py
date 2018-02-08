__author__='SeanPark_ViaSat'

import datetime
import time
import string
import random
import openpyxl
import pandas as pd
import xlsxwriter
import xlwt

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import selenium.webdriver.support.ui as ui
from selenium.webdriver.support.ui import Select

wb = openpyxl.load_workbook('NewConnectOrders.xlsx')

wbAddress = openpyxl.load_workbook(('./Data/Addresses.xlsx'))
sheetAddress = wbAddress['Sheet1']
username = sheetAddress.cell(row=4, column=2).value
password = sheetAddress.cell(row=4, column=3).value
salesChannel = sheetAddress.cell(row=4, column=4).value
customerType = sheetAddress.cell(row=4, column=5).value

driver=webdriver.Chrome("C:\\Selenium\\chromedriver.exe")
# driver = webdriver.Ie("C:\Selenium\IEDriverServer.exe");

driver.implicitly_wait(50)

driver.set_page_load_timeout(30)

driver.get("https://ordermgmt.test.exede.net/PublicGUI-SupportGUI/v1/pages/addcustomer/serviceAvailability.xhtml")

driver.maximize_window()

driver.implicitly_wait(20)

driver.find_element_by_xpath("//*[@id=\"document:body\"]/table/tbody/tr[2]/td/form/table/tbody/tr[3]/td[2]/input").send_keys(username)

driver.find_element_by_xpath("//*[@id=\"document:body\"]/table/tbody/tr[2]/td/form/table/tbody/tr[4]/td[2]/input").send_keys(password)

driver.find_element_by_name("submit").click()

driver.implicitly_wait(5)

addCustomerTab = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.XPATH, '//*[@id="add"]'))
)

addCustomerTab.click()

for item in range(3, 4):

    driver.implicitly_wait(3)
    time.sleep(1)

    if salesChannel == 'WB_DIRECT':
        driver.find_element_by_xpath("//*[@id=\"addCustomerForm:salesChannelMenu\"]/option[2]").click()

    # salesChannelOption = selectSalesChannel(salesChannel)
    # select = Select(driver.find_element_by_id('addCustomerForm:salesChannelMenu'))
    #
    # select.select_by_visible_text(salesChannel).click()

    now = datetime.datetime.now()

    currentYear = str(now.year)

    months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']

    hexdigits = list(string.hexdigits)
    del hexdigits[10:16]

    # print(hexdigits)

    randomMac = "AA:BB:CC:"

    for x in range(0, 6):
        randomNumber = random.choice(hexdigits)
        randomMac = randomMac + randomNumber
        if x % 2 != 0 and len(randomMac) < 17:
            randomMac = randomMac + ":"

    print("Mac Address : " + randomMac)

    randomMacNoColon = randomMac.replace(':', '')

    print(randomMacNoColon)

    currentMonth = months[now.month - 1]

    currentDay = ""

    if now.day < 10:
        currentDay = '0' + str(now.day)

    transactionReference = "SPark_" + currentDay + currentMonth + currentYear + str(item)

    newSheet = currentDay + "-" + currentMonth + "-" + currentYear

    wb.create_sheet(newSheet)

    ws = wb[newSheet]

    ws.cell(row=1, column=2).value = 'Transaction Reference'
    ws.cell(row=1, column=3).value = 'Service Agreement'
    ws.cell(row=1, column=4).value = 'MAC'
    ws.cell(row=2, column=2).value = transactionReference
    ws.cell(row=2, column=4).value = randomMac

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:transactionReference\"]").send_keys(transactionReference)

    # driver.implicitly_wait(2)
    # driver.find_element_by_xpath("//*[@id=\"addCustomerForm:namesIdName1\"]").send_keys("Spider")
    # time.sleep(1)

    firstNameField = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH,
                                        "//*[@id=\"addCustomerForm:namesIdName1\"]"))
    )

    firstNameField.send_keys("Spider")

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:namesIdName3\"]").send_keys("Man")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableAddress1\"]").send_keys("12017 E Lake Cir")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableCity\"]").send_keys("Englewood")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableStateAddressState\"]/option[7]").click()
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:addressIdMaybeTableZip\"]").send_keys("80111")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:primaryPhoneIdMaybeTablePhoneNumber\"]").send_keys("7204823823")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:emailAddressId\"]").send_keys("sean.park@viasat.com")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:Birthdate\"]").send_keys("12/15/1973")
    time.sleep(1)

    driver.implicitly_wait(2)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:nextButtonId\"]").click()
    time.sleep(1)

    # Contacts Page

    creditCheckPassed = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, '// *[@id = "addCustomerForm:_id91"]/tbody/tr/td/span'))
    )

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:customerReference\"]").send_keys(transactionReference)

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:accountReference\"]").send_keys(transactionReference)

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:nextButtonId\"]").click()

    # Packages Page

    packagesTitle = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:packagesHeaderLabel"]'))
    )

    # Sometimes, radio button is not checked by default. So intentionally click it.
    driver.find_element_by_xpath('//*[@id="addCustomerForm:_id112:_2"]').click()

    # "is not clickable at point" error. Another element is covering the element to click. I could use execute_script() to click on this.
    nextButton = driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]')
    driver.execute_script("arguments[0].click();", nextButton)

    driver.implicitly_wait(2)
    time.sleep(1)

    # Options Page
    optionsTitle = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:optionsLabel"]'))
    )

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:_1selectionPackages:_1\"]").click()

    driver.implicitly_wait(2)
    time.sleep(1)
    # driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]').click()

    optionsPageNextButton = driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]')
    driver.execute_script("arguments[0].click();", optionsPageNextButton)

    driver.implicitly_wait(2)
    time.sleep(1)

    # Payment Page
    paymentMethodTitle = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:recurringPaymentInfoLabel"]'))
    )

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdcreditCardTypeId\"]/option[3]").click()

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdNumberId\"]").send_keys("4012000077777777")

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath("//*[@id=\"addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdExpireMonthIdMonthId\"]/option[5]").click()

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdExpireYearIdYearId"]/option[3]').click()

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdFirstNameId"]').send_keys("VISA")

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdLastNameId"]').send_keys("APPROVAL")

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="addCustomerForm:recurringPaymentIdRecurringPaymentMethodIdTableCreditCardIdAddressZip"]').send_keys("80111")

    driver.implicitly_wait(2)
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="addCustomerForm:nextButtonId"]').click()

    # Review Page

    scheduleButton = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:scheduleInstallationButtonId"]'))
    )

    scheduleButton.click()
    time.sleep(1)

    # Schedule Page

    submitOrderButton = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:submitButtonId"]'))
    )

    submitOrderButton.click()
    time.sleep(0.3)

    # wait for order reference number created

    # Confirmation Page

    printButton = WebDriverWait(driver, 180).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:printButtonId"]'))
    )

    serviceAgreementReference = driver.find_element_by_xpath('//*[@id="addCustomerForm:serviceAgreementReference"]').text

    print("Sales Channel : " + salesChannel)
    print("External Account Reference : " + serviceAgreementReference)

    newOrderButton = driver.find_element_by_xpath('//*[@id="addCustomerForm:newOrderButtonId"]')

    newOrderButton.click()

    driver.implicitly_wait(20)
    time.sleep(1)

    transactionInfoTitle = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="addCustomerForm:transactionInfoLabel"]'))
    )

    driver.get('https://spyglass01.test.wdc1.wildblue.net:8443/SpyGlass/')

    referenceType = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[1]/select/option[5]'))
        )

    referenceType.click()

    referenceValue = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[2]/input'))
        )

    referenceValue.send_keys(transactionReference)

    externalSystem = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[3]/div/select/option[2]'))
        )

    externalSystem.click()

    driver.implicitly_wait(20)
    time.sleep(1)

    driver.find_element_by_xpath('/html/body/table/tbody/tr[2]/td/div/div/div[1]/div/div/form/div/table/tbody/tr/td[4]/input[1]').click()

    fsmCustomerCode = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="datatable"]/tbody/tr[1]/td[1]/div[1]'))
        )

    serviceAgreementNumber = driver.find_element_by_xpath('//*[@id="data"]/table[1]/tbody/tr[2]/td/table/tbody/tr[2]/td[12]').text

    print('serviceAgreementNumber : '+ serviceAgreementNumber)

    driver.save_screenshot('./Reports/'+serviceAgreementNumber+'.png')

    driver.implicitly_wait(20)
    time.sleep(1)

    ws.cell(row=2, column=3).value = serviceAgreementNumber

wb.save('NewConnectOrders.xlsx')

# time.sleep(10)

# save to the project home directory
# driver.get_screenshot_as_file(".\\Screenshots\\facebook.png");

# print(driver.title)

# assert "Facebook" in driver.title

# driver.find_element_by_id("email").send_keys("Selenium Webdriver")
#
# driver.find_element_by_name("pass").send_keys("Python")
#
# driver.find_element_by_id("loginbutton").click()
#
# driver.quit()

def selectSalesChannel(x):
    return {
        'WB_DIRECT': 2,
        'ATT': 3,
        'MEXICO_RETAIL': 4,
        'US_COMMUNITIES': 5,
        'US_SMALLBUSINESS':6,
        'DISH_DIRECT_RETAIL':7,
        'MEDIA_NETWORKS':8,
        'B2B_PARTNERS':9,
        'FIELD_TRIAL':10
    }.get(x, 'WB_DIRECT')